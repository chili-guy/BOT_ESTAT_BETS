#!/usr/bin/env python3
"""
Script para buscar estatísticas de jogadores da Premier League:
MINUTES, GOAL, ASSISTS, XG e XA para cada jogador em cada partida.

Permite buscar partidas por período específico definido pelo usuário.
"""

import pandas as pd
import requests
from bs4 import BeautifulSoup
import time
from datetime import datetime
import re
from urllib.parse import urljoin
import argparse
import sys

# Tentar importar cloudscraper para contornar proteções anti-bot
try:
    import cloudscraper
    HAS_CLOUDSCRAPER = True
except ImportError:
    HAS_CLOUDSCRAPER = False
    print("💡 Dica: Instale cloudscraper para melhor proteção anti-bot: pip install cloudscraper")


class PremierLeagueScraper:
    """Scraper para buscar dados da Premier League do fbref.com"""
    
    def __init__(self):
        self.base_url = "https://fbref.com"
        
        # Usar cloudscraper se disponível, senão usar requests normal
        if HAS_CLOUDSCRAPER:
            self.session = cloudscraper.create_scraper()
        else:
            self.session = requests.Session()
        
        # Headers para simular navegador real
        headers = {
            'User-Agent': 'Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
            'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8',
            'Accept-Language': 'en-US,en;q=0.9,pt-BR;q=0.8,pt;q=0.7',
            'Accept-Encoding': 'gzip, deflate, br',
            'DNT': '1',
            'Connection': 'keep-alive',
            'Upgrade-Insecure-Requests': '1',
            'Referer': 'https://www.google.com/'
        }
        
        if hasattr(self.session, 'headers'):
            self.session.headers.update(headers)
        
        self._initialized = False
    
    def _ensure_initialized(self):
        """Garante que a sessão foi inicializada"""
        if self._initialized:
            return
        
        try:
            if HAS_CLOUDSCRAPER:
                print("  ✅ Usando cloudscraper para contornar proteções anti-bot")
            else:
                print("  ⚠️  Usando requests padrão (pode ter problemas com proteções anti-bot)")
            
            print("  🔄 Estabelecendo conexão inicial...")
            initial_response = self.session.get(self.base_url, timeout=15)
            if initial_response.status_code == 200:
                print("  ✅ Conexão estabelecida com sucesso")
            else:
                print(f"  ⚠️  Resposta inicial: {initial_response.status_code}")
            self._initialized = True
        except Exception as e:
            print(f"  ⚠️  Aviso na conexão inicial: {e}")
            self._initialized = True
    
    def get_player_stats_from_match(self, match_url, team, opponent, date, location):
        """
        Extrai estatísticas de jogadores de um jogo específico.
        Retorna lista de dicionários com: Player, Team, Date, Opponent, Minutes, Goals, Assists, xG, xA
        """
        try:
            if not match_url or '/matches/' not in match_url:
                print(f"    ⚠️  URL inválida: {match_url}")
                return []
            
            print(f"    🔗 Acessando: {match_url}")
            time.sleep(3)  # Rate limiting
            
            response = self.session.get(match_url, timeout=20)
            
            if response.status_code == 429:
                print(f"    ⚠️  Rate limit (429). Aguardando 30 segundos...")
                time.sleep(30)
                response = self.session.get(match_url, timeout=20)
            
            response.raise_for_status()
            soup = BeautifulSoup(response.content, 'html.parser')
            
            # Verificar se é página de jogo válida ou página genérica
            page_title = soup.find('title')
            if page_title:
                title_text = page_title.get_text().lower()
                if 'schedule' in title_text or 'fixtures' in title_text:
                    print(f"    ⚠️  Página parece ser de schedule, não de jogo individual")
                    print(f"    💡 URL pode estar incorreta: {match_url}")
                    return []
            
            player_stats = []
            
            # Debug: listar todas as tabelas encontradas
            all_tables = soup.find_all('table', {'id': re.compile(r'.*')})
            table_ids = [t.get('id', 'N/A') for t in all_tables if t.get('id')]
            
            # Filtrar tabelas relevantes
            relevant_tables = [tid for tid in table_ids if 'sched' not in tid.lower()]
            
            if relevant_tables:
                print(f"    🔍 Tabelas encontradas: {len(relevant_tables)} (primeiras 5: {relevant_tables[:5]})")
            else:
                print(f"    ⚠️  Nenhuma tabela relevante encontrada (todas as tabelas: {table_ids[:5]})")
            
            # Encontrar tabela de estatísticas do time (home ou away)
            stats_table = None
            
            # Padrões para encontrar a tabela correta
            # Nota: fbref usa IDs únicos (ex: stats_4ba7cbea_summary), não stats_home_summary
            # Precisamos buscar por tabelas que contenham "summary" no ID
            
            # Primeiro, buscar todas as tabelas summary disponíveis
            summary_tables = []
            for table in all_tables:
                table_id = table.get('id', '').lower()
                # Buscar tabelas que contenham "summary" no ID (são as tabelas principais de estatísticas)
                if 'summary' in table_id and 'stats' in table_id:
                    summary_tables.append(table)
            
            # Se encontrou tabelas summary, usar baseado na posição (home geralmente é primeira, away é segunda)
            if summary_tables:
                if location == 'home' and len(summary_tables) > 0:
                    stats_table = summary_tables[0]
                    print(f"    ✓ Usando primeira tabela summary para home team")
                elif location == 'away' and len(summary_tables) > 1:
                    stats_table = summary_tables[1]
                    print(f"    ✓ Usando segunda tabela summary para away team")
                elif len(summary_tables) == 1:
                    stats_table = summary_tables[0]
                    print(f"    ✓ Usando única tabela summary disponível")
            
            # Se não encontrou summary, tentar padrões antigos como fallback
            if not stats_table:
                patterns = [
                    f'stats_{location}_summary',
                    f'stats_{location}_players',
                    f'stats_{location}',
                ]
                
                for pattern in patterns:
                    stats_table = soup.find('table', {'id': pattern})
                    if stats_table:
                        break
            
            # Método alternativo: procurar por tabelas com estrutura de estatísticas de jogadores
            if not stats_table:
                for table in all_tables:
                    table_id = table.get('id', '').lower()
                    if 'stats' in table_id and 'summary' in table_id:
                        thead = table.find('thead')
                        if thead:
                            headers = [th.get_text(strip=True).lower() for th in thead.find_all('th')]
                            header_text = ' '.join(headers)
                            if 'player' in header_text and ('min' in header_text or 'goals' in header_text):
                                stats_table = table
                                break
            
            # Último fallback: usar qualquer tabela com player (não recomendado, mas melhor que nada)
            if not stats_table:
                candidate_tables = []
                for table in all_tables:
                    thead = table.find('thead')
                    if thead:
                        headers = [th.get_text(strip=True).lower() for th in thead.find_all('th')]
                        header_text = ' '.join(headers)
                        if ('player' in header_text and 
                            ('min' in header_text or 'goals' in header_text or 'assists' in header_text)):
                            candidate_tables.append(table)
                
                # Filtrar para pegar apenas summary se possível
                summary_candidates = [t for t in candidate_tables if 'summary' in t.get('id', '').lower()]
                if summary_candidates:
                    candidate_tables = summary_candidates
                
                # Home geralmente é primeira tabela, away é segunda
                if candidate_tables:
                    if location == 'home' and len(candidate_tables) > 0:
                        stats_table = candidate_tables[0]
                    elif location == 'away' and len(candidate_tables) > 1:
                        stats_table = candidate_tables[1]
                    elif len(candidate_tables) == 1:
                        stats_table = candidate_tables[0]
            
            if not stats_table:
                print(f"    ❌ Tabela de estatísticas não encontrada para {location} team")
                print(f"    💡 Tabelas disponíveis: {relevant_tables[:10]}")
                # Tentar uma última vez: procurar qualquer tabela com "player" no cabeçalho
                for table in all_tables:
                    thead = table.find('thead')
                    if thead:
                        header_text = ' '.join([th.get_text(strip=True).lower() for th in thead.find_all('th')])
                        if 'player' in header_text:
                            # Verificar se tem colunas de estatísticas
                            if any(word in header_text for word in ['min', 'goals', 'assists', 'xg', 'xa']):
                                print(f"    💡 Tentando usar tabela genérica: {table.get('id', 'N/A')}")
                                stats_table = table
                                break
                
                if not stats_table:
                    # Verificar se o jogo foi realmente jogado (procurar por placar ou resultado)
                    score_elements = soup.find_all(['div', 'span'], string=re.compile(r'\d+\s*-\s*\d+'))
                    if not score_elements:
                        print(f"    ℹ️  Jogo pode não ter sido jogado ainda - sem placar visível")
                    return []
            
            print(f"    ✓ Tabela encontrada ({stats_table.get('id', 'N/A')})")
            
            # Debug apenas se necessário (comentado para produção)
            # header_row = stats_table.find('thead')
            # if header_row:
            #     all_headers = [th.get_text(strip=True) for th in header_row.find_all('th')]
            #     all_data_stats = [th.get('data-stat', 'N/A') for th in header_row.find_all('th')]
            #     # Procurar xA/xAG nas colunas
            #     xa_found = False
            #     for i, (header, data_stat) in enumerate(zip(all_headers, all_data_stats)):
            #         if 'xa' in header.lower() or 'xag' in header.lower() or 'xa' in data_stat.lower():
            #             xa_found = True
            #             print(f"    🔍 DEBUG: Coluna {i} - Header: '{header}' | data-stat: '{data_stat}'")
            #     if not xa_found:
            #         print(f"    ⚠️  DEBUG: xA não encontrado nos cabeçalhos. Colunas disponíveis:")
            #         for i, (header, data_stat) in enumerate(zip(all_headers[:15], all_data_stats[:15])):
            #             print(f"        {i}: '{header}' ({data_stat})")
            
            rows = stats_table.find_all('tr')[1:]  # Pular cabeçalho
            
            first_row_processed = False
            for row in rows:
                # Pular linhas de subtotais e cabeçalhos
                row_class = str(row.get('class', []))
                if 'thead' in row_class or 'spacer' in row_class:
                    continue
                
                cells = row.find_all(['td', 'th'])
                if len(cells) < 3:
                    continue
                
                # Extrair dados usando data-stat (método mais confiável no fbref)
                player_name = ""
                minutes = 0
                goals = 0
                assists = 0
                xg = 0.0
                xa = 0.0
                
                # Primeiro loop: buscar todos os valores básicos
                for cell in cells:
                    data_stat = cell.get('data-stat', '').lower()
                    text = cell.get_text(strip=True)
                    
                    if data_stat == 'player':
                        player_name = text
                    elif data_stat == 'minutes':
                        try:
                            minutes = int(re.sub(r'[^\d]', '', text) or 0)
                        except:
                            minutes = 0
                    elif data_stat == 'goals':
                        try:
                            goals = int(re.sub(r'[^\d]', '', text) or 0)
                        except:
                            goals = 0
                    elif data_stat == 'assists':
                        try:
                            assists = int(re.sub(r'[^\d]', '', text) or 0)
                        except:
                            assists = 0
                    elif data_stat == 'xg':
                        try:
                            # Converter vírgula para ponto (formato brasileiro/europeu)
                            text_clean = text.replace(',', '.')
                            # Remover tudo exceto dígitos e ponto
                            text_clean = re.sub(r'[^\d.]', '', text_clean)
                            if text_clean:
                                xg = float(text_clean)
                            else:
                                xg = 0.0
                        except (ValueError, AttributeError):
                            xg = 0.0
                
                # Buscar xA APENAS usando xg_assist (nome correto no fbref)
                # NÃO usar fallbacks que podem pegar valores errados
                xa = 0.0
                for cell in cells:
                    data_stat = cell.get('data-stat', '').lower()
                    
                    # APENAS usar xg_assist - método mais confiável
                    if data_stat == 'xg_assist':
                        try:
                            text = cell.get_text(strip=True)
                            # Converter vírgula para ponto
                            text_clean = text.replace(',', '.')
                            # Remover tudo exceto dígitos e ponto
                            text_clean = re.sub(r'[^\d.]', '', text_clean)
                            if text_clean:
                                xa = float(text_clean)
                            else:
                                xa = 0.0
                            break  # Parar assim que encontrar (só existe uma célula xg_assist)
                        except (ValueError, AttributeError, TypeError):
                            xa = 0.0
                            break
                
                # Fallback: se não encontrou por data-stat, tentar por posição
                if not player_name and len(cells) > 0:
                    player_name = cells[0].get_text(strip=True)
                
                # Ignorar linhas sem nome válido ou linhas de subtotal/agregado
                if not player_name or player_name in ['Player', '', 'Reserves', 'Team Total']:
                    continue
                
                # Filtrar linhas de subtotal/agregado
                # Padrões comuns: "16 Players", "15 Players", etc.
                player_name_lower = player_name.lower()
                is_subtotal = False
                
                # Verificar se o nome parece ser um subtotal (ex: "16 Players", "15 Players")
                if 'player' in player_name_lower and any(char.isdigit() for char in player_name):
                    is_subtotal = True
                
                # Verificar se tem minutos anormalmente altos (subtotais)
                # Um jogo tem no máximo ~120 minutos (90 + tempo extra)
                # Subtotais geralmente têm 990 minutos (11 jogadores x 90 minutos)
                if minutes > 120:
                    is_subtotal = True
                
                # Verificar padrões específicos de subtotal
                if re.match(r'^\d+\s+[Pp]layers?$', player_name.strip()):
                    is_subtotal = True
                
                if is_subtotal:
                    continue
                
                # Fallback: Se não encontrou estatísticas básicas, tentar por índice no cabeçalho
                # MAS NÃO buscar xA por índice - xA já foi buscado usando xg_assist acima
                if minutes == 0 and goals == 0 and assists == 0:
                    header_row = stats_table.find('thead')
                    if header_row:
                        headers = [th.get_text(strip=True) for th in header_row.find_all('th')]
                        try:
                            min_idx = next((i for i, h in enumerate(headers) if 'min' in h.lower()), -1)
                            gls_idx = next((i for i, h in enumerate(headers) if 'gls' in h.lower() or h.lower() == 'g'), -1)
                            ast_idx = next((i for i, h in enumerate(headers) if 'ast' in h.lower() or h.lower() == 'a'), -1)
                            xg_idx = next((i for i, h in enumerate(headers) if 'xg' in h.lower() and 'xag' not in h.lower()), -1)
                            
                            if min_idx >= 0 and min_idx < len(cells):
                                minutes_str = cells[min_idx].get_text(strip=True)
                                minutes = int(re.sub(r'[^\d]', '', minutes_str) or 0)
                            
                            if gls_idx >= 0 and gls_idx < len(cells):
                                goals = int(re.sub(r'[^\d]', '', cells[gls_idx].get_text(strip=True)) or 0)
                            
                            if ast_idx >= 0 and ast_idx < len(cells):
                                assists = int(re.sub(r'[^\d]', '', cells[ast_idx].get_text(strip=True)) or 0)
                            
                            if xg_idx >= 0 and xg_idx < len(cells):
                                xg_str = cells[xg_idx].get_text(strip=True)
                                try:
                                    # Converter vírgula para ponto (formato brasileiro/europeu)
                                    xg_str_clean = xg_str.replace(',', '.')
                                    xg_str_clean = re.sub(r'[^\d.]', '', xg_str_clean)
                                    if xg_str_clean:
                                        xg = float(xg_str_clean)
                                    else:
                                        xg = 0.0
                                except (ValueError, AttributeError):
                                    xg = 0.0
                            
                            # xA NÃO é buscado aqui - já foi buscado usando xg_assist no loop acima
                            # Isso garante que sempre usamos o valor correto de xg_assist
                        except (ValueError, IndexError):
                            pass
                
                # xA já foi extraído usando xg_assist no loop acima
                # NÃO usar métodos alternativos/fallbacks que podem pegar valores errados
                # Se xa == 0.0, pode ser que realmente seja 0
                
                # Format xG e xA com 4 casas decimais (garantir que sempre mostra 4 decimais)
                xg_formatted = round(float(xg), 4) if xg > 0 else 0.0000
                xa_formatted = round(float(xa), 4) if xa > 0 else 0.0000
                
                # Criar registro
                stats = {
                    'Player': player_name,
                    'Team': team,
                    'Date': date,
                    'Opponent': opponent,
                    'Minutes': minutes,
                    'Goals': goals,
                    'Assists': assists,
                    'xG': xg_formatted,
                    'xA': xa_formatted,
                    'Confronto': f"{team}|{opponent}|{date.strftime('%Y-%m-%d')}",
                    'Location': location,
                    'adj': 0,
                    'Year': date.year,
                    'Month': date.month
                }
                player_stats.append(stats)
            
            print(f"    ✓ {len(player_stats)} jogadores processados")
            return player_stats
            
        except requests.exceptions.HTTPError as e:
            if e.response.status_code == 429:
                print(f"    ⚠️  Rate limit (429)")
                time.sleep(60)
            else:
                print(f"    ❌ Erro HTTP {e.response.status_code}: {e}")
            return []
        except Exception as e:
            print(f"    ❌ Erro ao extrair estatísticas: {e}")
            return []


def process_schedule_table(soup, start_date, end_date, scraper, limit_games=None):
    """
    Processa a tabela de jogos e extrai estatísticas de jogadores
    para jogos dentro do período especificado
    """
    all_player_stats = []
    
    # Encontrar tabela de jogos
    table = soup.find('table', {'id': 'sched_9_1'})
    if not table:
        table = soup.find('table', {'id': re.compile(r'sched.*')})
    if not table:
        tables = soup.find_all('table')
        for t in tables:
            if 'schedule' in str(t.get('id', '')).lower() or 'fixture' in str(t.get('id', '')).lower():
                table = t
                break
    
    if not table:
        print(f"  ⚠️  Tabela de jogos não encontrada")
        return []
    
    rows = table.find_all('tr')[1:]  # Pular cabeçalho
    print(f"  Encontradas {len(rows)} linhas na tabela")
    
    matches_found = 0
    skipped_no_date = 0
    skipped_out_of_range = 0
    skipped_no_teams = 0
    
    for row_idx, row in enumerate(rows):
        cells = row.find_all(['td', 'th'])
        if len(cells) < 3:
            continue
        
        try:
            # Extrair data
            date_text = ""
            date_cell = None
            
            # Procurar por atributo data-date
            for cell in cells:
                date_attr = cell.get('data-date', '')
                if date_attr and re.match(r'\d{4}-\d{2}-\d{2}', date_attr):
                    date_text = date_attr
                    date_cell = cell
                    break
            
            # Se não encontrou, procurar por texto
            if not date_text:
                for cell in cells[:10]:
                    text = cell.get_text(strip=True)
                    if re.match(r'^\d{4}-\d{2}-\d{2}$', text):
                        date_text = text
                        date_cell = cell
                        break
            
            if not date_text:
                skipped_no_date += 1
                continue
            
            # Converter data
            match_date = pd.to_datetime(date_text, format='%Y-%m-%d', errors='coerce')
            if pd.isna(match_date):
                skipped_no_date += 1
                continue
            
            # Verificar se está no período especificado
            if match_date < start_date or match_date > end_date:
                skipped_out_of_range += 1
                continue
            
            # Extrair times
            home_team = ""
            away_team = ""
            
            # Por data-stat
            for cell in cells:
                data_stat = cell.get('data-stat', '')
                text = cell.get_text(strip=True)
                
                if data_stat == 'home_team':
                    home_team = text
                elif data_stat == 'away_team':
                    away_team = text
            
            # Por posição padrão
            if not home_team and len(cells) > 4:
                home_team = cells[4].get_text(strip=True)
            if not away_team and len(cells) > 5:
                away_team = cells[5].get_text(strip=True)
            
            # Por links
            if not home_team or not away_team:
                team_links = []
                for cell in cells:
                    link = cell.find('a')
                    if link:
                        text = link.get_text(strip=True)
                        href = link.get('href', '')
                        if text and len(text) > 2 and '/squads/' in href:
                            if (not re.match(r'^\d+$', text) and 
                                not re.match(r'\d{4}-\d{2}-\d{2}', text) and
                                not re.match(r'\d{2}:\d{2}', text) and
                                text not in ['Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat', 'Sun']):
                                team_links.append(text)
                
                if team_links:
                    if not home_team and len(team_links) > 0:
                        home_team = team_links[0]
                    if not away_team and len(team_links) > 1:
                        away_team = team_links[1]
            
            if not home_team or not away_team:
                skipped_no_teams += 1
                continue
            
            # Verificar se o jogo foi jogado (tem placar) antes de buscar link
            # Jogos futuros não têm placar e não têm dados disponíveis
            has_score = False
            score_text = ""
            for cell in cells:
                text = cell.get_text(strip=True)
                # Verificar se tem placar (formato XX-XX ou XX:XX)
                if re.match(r'^\d+[\s\-:]\d+$', text) and len(text) <= 10:
                    # Verificar se não é hora (HH:MM)
                    if ':' not in text or (':' in text and int(text.split(':')[0]) > 23):
                        has_score = True
                        score_text = text
                        break
            
            # Se não tem placar, o jogo provavelmente não foi jogado ainda
            if not has_score:
                # Verificar também data-stat para score
                for cell in cells:
                    data_stat = cell.get('data-stat', '')
                    if 'score' in data_stat.lower() and cell.get_text(strip=True):
                        has_score = True
                        score_text = cell.get_text(strip=True)
                        break
                
                if not has_score:
                    print(f"  📅 {match_date.strftime('%Y-%m-%d')}: {home_team} vs {away_team} - ⏳ Jogo ainda não foi jogado (sem placar)")
                    continue
            
            # Buscar link do jogo - procurar em todas as células
            match_link = None
            
            # Primeiro, tentar na célula da data
            if date_cell:
                link_cell = date_cell.find('a')
                if link_cell and link_cell.get('href'):
                    href = link_cell.get('href', '')
                    if '/matches/' in href and '/schedule/' not in href:
                        match_link = urljoin(scraper.base_url, href)
                        if not match_link.startswith('http'):
                            match_link = urljoin(scraper.base_url, href)
            
            # Se não encontrou, procurar em todas as células
            if not match_link:
                for cell in cells:
                    links = cell.find_all('a')
                    for link in links:
                        href = link.get('href', '')
                        if href and '/matches/' in href and '/schedule/' not in href:
                            # Verificar se não é apenas uma data
                            if not re.match(r'^/?en/matches/\d{4}-\d{2}-\d{2}$', href):
                                match_link = urljoin(scraper.base_url, href)
                                if match_link.startswith('http'):
                                    break
                    if match_link:
                        break
            
            # Se ainda não encontrou link, mas tem placar, pode ser problema na extração
            if not match_link:
                print(f"     ⚠️  Link do jogo não encontrado (mas jogo foi jogado - placar: {score_text})")
                # Tentar métodos alternativos de busca de link
                # Às vezes o link está em outra célula ou formato diferente
                for cell in cells:
                    # Procurar todos os links na célula
                    all_links = cell.find_all('a')
                    for link in all_links:
                        href = link.get('href', '')
                        if href and ('/match' in href.lower() or 'match' in href.lower()):
                            if '/schedule' not in href.lower():
                                match_link = urljoin(scraper.base_url, href)
                                if match_link.startswith('http'):
                                    break
                    if match_link:
                        break
                
                if not match_link:
                    print(f"     ❌ Não foi possível encontrar link válido - pulando este jogo")
                    continue
            
            # Validar URL encontrada
            if match_link and not match_link.startswith('http'):
                match_link = urljoin(scraper.base_url, match_link)
            
            # Verificar se URL é válida (deve conter /matches/ e um ID válido)
            if match_link and '/matches/' in match_link:
                # Verificar se não é apenas data (formato inválido)
                if re.match(r'.*/matches/\d{4}-\d{2}-\d{2}$', match_link):
                    print(f"     ⚠️  URL parece inválida (apenas data): {match_link}")
                    print(f"     ℹ️  Tentando encontrar link correto...")
                    # Tentar novamente com método mais agressivo
                    match_link = None
                    for cell in cells:
                        all_links = cell.find_all('a')
                        for link in all_links:
                            href = link.get('href', '')
                            if href and '/matches/' in href:
                                # Verificar se tem ID único (geralmente tem letras/ID após data)
                                if not re.match(r'^/en/matches/\d{4}-\d{2}-\d{2}$', href):
                                    match_link = urljoin(scraper.base_url, href)
                                    if match_link.startswith('http'):
                                        break
                        if match_link:
                            break
                    
                    if not match_link:
                        print(f"     ❌ Não foi possível encontrar link válido - pulando")
                        continue
            
            # Verificar limite
            if limit_games and matches_found >= limit_games:
                print(f"  ⏸️  Limite de {limit_games} jogos atingido")
                break
            
            print(f"  📅 {match_date.strftime('%Y-%m-%d')}: {home_team} vs {away_team} (Placar: {score_text})")
            print(f"     🔗 URL: {match_link}")
            
            # Buscar estatísticas
            stats_home = scraper.get_player_stats_from_match(
                match_link, home_team, away_team, match_date, 'home'
            )
            all_player_stats.extend(stats_home)
            
            stats_away = scraper.get_player_stats_from_match(
                match_link, away_team, home_team, match_date, 'away'
            )
            all_player_stats.extend(stats_away)
            
            matches_found += 1
            print(f"     ✓ {len(stats_home) + len(stats_away)} jogadores processados")
            
            # Se não encontrou dados, pode ser jogo futuro
            if len(stats_home) == 0 and len(stats_away) == 0:
                print(f"     ⚠️  Nenhum dado encontrado - jogo pode não ter sido jogado ainda")
            
            time.sleep(5)  # Delay entre jogos
            
        except Exception as e:
            print(f"     ❌ Erro ao processar linha: {e}")
            continue
    
    print(f"\n  📊 Estatísticas:")
    print(f"     ✓ Jogos processados: {matches_found}")
    print(f"     ⚠️  Sem data: {skipped_no_date}")
    print(f"     ⚠️  Fora do período: {skipped_out_of_range}")
    print(f"     ⚠️  Sem times: {skipped_no_teams}")
    print(f"  ✅ Total: {len(all_player_stats)} registros de jogadores")
    
    return all_player_stats


def get_season_url(year, month):
    """Determina a URL da temporada baseado no mês"""
    # Temporada da Premier League vai de agosto a maio
    # Se o mês é agosto-dezembro, está na temporada atual ano-ano+1
    # Se o mês é janeiro-maio, está na temporada anterior ano-1-ano
    if month >= 8:
        season = f"{year}-{year+1}"
    else:
        season = f"{year-1}-{year}"
    
    return f"https://fbref.com/en/comps/9/{season}/schedule/{season}-Premier-League-Scores-and-Fixtures"


def scrape_period(start_date, end_date, scraper=None, limit_games=None):
    """Busca dados para um período específico"""
    if scraper is None:
        scraper = PremierLeagueScraper()
    
    if not scraper._initialized:
        scraper._ensure_initialized()
    
    all_player_stats = []
    
    # Gerar lista de meses para processar
    current_date = start_date
    months_to_process = []
    
    while current_date <= end_date:
        year_month = (current_date.year, current_date.month)
        if year_month not in months_to_process:
            months_to_process.append(year_month)
        
        # Avançar para o próximo mês
        if current_date.month == 12:
            current_date = current_date.replace(year=current_date.year + 1, month=1, day=1)
        else:
            current_date = current_date.replace(month=current_date.month + 1, day=1)
    
    print(f"\n📅 Período: {start_date.strftime('%Y-%m-%d')} até {end_date.strftime('%Y-%m-%d')}")
    print(f"📋 Meses a processar: {', '.join([f'{y}-{m:02d}' for y, m in months_to_process])}")
    
    for year, month in months_to_process:
        print(f"\n{'='*60}")
        print(f"Processando {year}-{month:02d}...")
        print(f"{'='*60}")
        
        try:
            url = get_season_url(year, month)
            print(f"  Acessando: {url}")
            
            response = scraper.session.get(url, timeout=15)
            response.raise_for_status()
            soup = BeautifulSoup(response.content, 'html.parser')
            
            # Processar tabela filtrando pelo período
            month_stats = process_schedule_table(soup, start_date, end_date, scraper, limit_games)
            all_player_stats.extend(month_stats)
            
            # Delay entre meses
            if (year, month) != months_to_process[-1]:
                print(f"  ⏳ Aguardando 10 segundos antes do próximo mês...")
                time.sleep(10)
                
        except Exception as e:
            print(f"  ❌ Erro ao processar {year}-{month:02d}: {e}")
            continue
    
    return all_player_stats


def get_spreadsheet_template(filepath='premier.xlsx'):
    """Obtém a estrutura de colunas da planilha existente"""
    try:
        df_template = pd.read_excel(filepath, nrows=0)
        return df_template.columns.tolist()
    except:
        # Estrutura padrão baseada na análise
        return [
            'Player', 'Team', 'Date', 'Opponent', 'Minutes', 'Goals', 'Assists',
            'xG', 'xA', 'Confronto', 'Location', 'adj', 'TEAM', 'Chelsea',
            'Unnamed: 14', 'LINHA', '1.5', 'Unnamed: 17', 'FAIR ASS',
            '6.83060124935948', 'Unnamed: 20', 'FAIR GOAL', '5.714158363009968',
            'Unnamed: 23', 'LOCAL TEAM', 'Unnamed: 25', 'Unnamed: 26',
            'LOCAL PLAYER', 'Unnamed: 28', 'Unnamed: 29', 'CONFRONTO',
            'Unnamed: 31', 'Unnamed: 32', 'Unnamed: 33', 'Neto', 'Arsenal',
            'Year', 'Month'
        ]


def main():
    parser = argparse.ArgumentParser(
        description='Busca estatísticas de jogadores da Premier League (MINUTES, GOALS, ASSISTS, XG, XA)',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Exemplos:
  # Buscar dados de outubro de 2025
  python buscar_estatisticas.py --inicio 2025-10-01 --fim 2025-10-31

  # Buscar dados de outubro a dezembro de 2025
  python buscar_estatisticas.py --inicio 2025-10-01 --fim 2025-12-31

  # Modo teste (limita a 5 jogos)
  python buscar_estatisticas.py --inicio 2025-10-01 --fim 2025-10-31 --limit 5 --test

  # Especificar arquivo de saída
  python buscar_estatisticas.py --inicio 2025-10-01 --fim 2025-10-31 --output minha_planilha.xlsx
        """
    )
    
    parser.add_argument(
        '--inicio',
        type=str,
        required=True,
        help='Data de início (formato: YYYY-MM-DD, ex: 2025-10-01)'
    )
    parser.add_argument(
        '--fim',
        type=str,
        required=True,
        help='Data de fim (formato: YYYY-MM-DD, ex: 2025-10-31)'
    )
    parser.add_argument(
        '--output',
        type=str,
        default='premier.xlsx',
        help='Arquivo Excel de saída (padrão: premier.xlsx)'
    )
    parser.add_argument(
        '--limit',
        type=int,
        default=None,
        help='Limitar número de jogos a processar (útil para testes)'
    )
    parser.add_argument(
        '--test',
        action='store_true',
        help='Modo teste - não salva arquivo, apenas mostra resultados'
    )
    
    args = parser.parse_args()
    
    print("="*60)
    print("BUSCADOR DE ESTATÍSTICAS - PREMIER LEAGUE")
    print("="*60)
    print("\nEstatísticas buscadas: MINUTES, GOALS, ASSISTS, XG, XA")
    
    if args.test:
        print("🧪 MODO TESTE - Nenhuma alteração será salva")
    
    # Parsear datas
    try:
        start_date = pd.to_datetime(args.inicio)
        end_date = pd.to_datetime(args.fim)
    except:
        print("❌ Erro: Formato de data inválido. Use YYYY-MM-DD (ex: 2025-10-01)")
        return
    
    if start_date > end_date:
        print("❌ Erro: Data de início deve ser anterior à data de fim")
        return
    
    print(f"\n📅 Período: {start_date.strftime('%Y-%m-%d')} até {end_date.strftime('%Y-%m-%d')}")
    
    if not args.test:
        resposta = input("\n⚠️  Continuar com a busca? (s/n): ").strip().lower()
        if resposta not in ['s', 'sim', 'y', 'yes']:
            print("Operação cancelada.")
            return
    
    # Inicializar scraper
    print("\n🔧 Inicializando scraper...")
    scraper = PremierLeagueScraper()
    scraper._ensure_initialized()
    
    # Buscar dados
    print("\n🚀 Iniciando busca...")
    all_data = scrape_period(start_date, end_date, scraper, limit_games=args.limit)
    
    if not all_data:
        print("\n⚠️  Nenhum dado foi encontrado.")
        print("\n💡 Possíveis razões:")
        print("   - Os dados ainda não estão disponíveis no site")
        print("   - A estrutura do site mudou")
        print("   - Problemas de conexão ou bloqueio")
        return
    
    # Converter para DataFrame
    print(f"\n📊 Processando {len(all_data)} registros coletados...")
    new_df = pd.DataFrame(all_data)
    
    # Obter estrutura da planilha original
    try:
        template_columns = get_spreadsheet_template(args.output)
        print(f"  📋 Usando estrutura da planilha existente ({len(template_columns)} colunas)")
    except:
        template_columns = get_spreadsheet_template('premier.xlsx')
        print(f"  📋 Usando estrutura padrão ({len(template_columns)} colunas)")
    
    # Garantir que todas as colunas existam
    for col in template_columns:
        if col not in new_df.columns:
            new_df[col] = None
    
    # Reordenar colunas
    new_df = new_df[template_columns]
    
    # Remover duplicatas
    before_dedup = len(new_df)
    new_df = new_df.drop_duplicates(subset=['Player', 'Team', 'Date', 'Opponent'], keep='last')
    duplicates_removed = before_dedup - len(new_df)
    
    if duplicates_removed > 0:
        print(f"  🧹 Removidas {duplicates_removed} duplicatas")
    
    # Ordenar por data
    new_df = new_df.sort_values('Date')
    
    if args.test:
        print(f"\n{'='*60}")
        print("🧪 MODO TESTE - Resultados:")
        print(f"{'='*60}")
        print(f"📈 Total de registros: {len(new_df)}")
        print(f"📅 Data mínima: {new_df['Date'].min().strftime('%Y-%m-%d')}")
        print(f"📅 Data máxima: {new_df['Date'].max().strftime('%Y-%m-%d')}")
        print(f"\nPrimeiras 5 linhas:")
        print(new_df.head(5).to_string())
        print(f"\n💡 Execute sem --test para salvar a planilha")
        print(f"{'='*60}")
    else:
        # Salvar planilha
        import os
        from pathlib import Path
        from openpyxl import load_workbook
        from openpyxl.styles import NamedStyle
        
        # Criar diretório se não existir
        output_path = Path(args.output)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        
        # Salvar DataFrame no Excel
        new_df.to_excel(args.output, index=False)
        
        # Formatar colunas xG e xA com 4 casas decimais
        try:
            wb = load_workbook(args.output)
            ws = wb.active
            
            # Encontrar índices das colunas xG e xA
            header_row = 1
            xg_col = None
            xa_col = None
            
            for col_idx, cell in enumerate(ws[header_row], 1):
                if cell.value == 'xG':
                    xg_col = col_idx
                elif cell.value == 'xA':
                    xa_col = col_idx
            
            # Aplicar formatação de 4 casas decimais
            from openpyxl.styles import NamedStyle
            decimal_style = NamedStyle(name="decimal4", number_format="0.0000")
            
            if xg_col:
                for row in range(2, ws.max_row + 1):
                    cell = ws.cell(row=row, column=xg_col)
                    if cell.value is not None:
                        try:
                            cell.value = float(cell.value)
                            cell.number_format = "0.0000"
                        except (ValueError, TypeError):
                            pass
            
            if xa_col:
                for row in range(2, ws.max_row + 1):
                    cell = ws.cell(row=row, column=xa_col)
                    if cell.value is not None:
                        try:
                            cell.value = float(cell.value)
                            cell.number_format = "0.0000"
                        except (ValueError, TypeError):
                            pass
            
            # Salvar workbook
            wb.save(args.output)
            wb.close()
        except Exception as e:
            print(f"  ⚠️  Aviso: Não foi possível formatar decimais: {e}")
            print(f"  💡 Os valores serão salvos, mas podem não mostrar 4 casas decimais fixas")
        
        print(f"\n{'='*60}")
        print("✅ PLANILHA SALVA COM SUCESSO!")
        print(f"{'='*60}")
        print(f"📈 Total de registros: {len(new_df)}")
        print(f"📅 Data mínima: {new_df['Date'].min().strftime('%Y-%m-%d')}")
        print(f"📅 Data máxima: {new_df['Date'].max().strftime('%Y-%m-%d')}")
        print(f"📁 Arquivo salvo: {os.path.abspath(args.output)}")
        print(f"{'='*60}")


if __name__ == "__main__":
    main()

