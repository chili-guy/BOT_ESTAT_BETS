#!/usr/bin/env python3
"""
Script genérico para buscar estatísticas de jogadores de múltiplas ligas:
- La Liga (Espanha)
- Bundesliga (Alemanha)
- Serie A (Itália)
- Primeira Liga (Portugal)
- Ligue 1 (França)
- Championship (Inglaterra - Série B)

Estatísticas: MINUTES, GOALS, ASSISTS, XG e XA
"""

import pandas as pd
import requests
from bs4 import BeautifulSoup
import time
from datetime import datetime
import re
from urllib.parse import urljoin
import argparse
import sys

# Tentar importar cloudscraper
try:
    import cloudscraper
    HAS_CLOUDSCRAPER = True
except ImportError:
    HAS_CLOUDSCRAPER = False

# IDs das ligas no fbref.com
LEAGUE_IDS = {
    'premier': {'id': 9, 'name': 'Premier League', 'country': 'Inglaterra'},
    'laliga': {'id': 12, 'name': 'La Liga', 'country': 'Espanha'},
    'bundesliga': {'id': 20, 'name': 'Bundesliga', 'country': 'Alemanha'},
    'seriea': {'id': 11, 'name': 'Serie A', 'country': 'Itália'},
    'portugal': {'id': 32, 'name': 'Primeira Liga', 'country': 'Portugal'},
    'ligue1': {'id': 13, 'name': 'Ligue 1', 'country': 'França'},
    'championship': {'id': 10, 'name': 'Championship', 'country': 'Inglaterra (Série B)'},
}

class LeagueScraper:
    """Scraper genérico para buscar dados de qualquer liga do fbref.com"""
    
    def __init__(self):
        self.base_url = "https://fbref.com"
        
        if HAS_CLOUDSCRAPER:
            self.session = cloudscraper.create_scraper()
        else:
            self.session = requests.Session()
        
        headers = {
            'User-Agent': 'Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
            'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8',
            'Accept-Language': 'en-US,en;q=0.9,pt-BR;q=0.8,pt;q=0.7',
            'Accept-Encoding': 'gzip, deflate, br',
            'DNT': '1',
            'Connection': 'keep-alive',
            'Upgrade-Insecure-Requests': '1',
            'Referer': 'https://www.google.com/'
        }
        
        if hasattr(self.session, 'headers'):
            self.session.headers.update(headers)
        
        self._initialized = False
    
    def _ensure_initialized(self):
        """Garante que a sessão foi inicializada"""
        if self._initialized:
            return
        
        try:
            if HAS_CLOUDSCRAPER:
                print("  ✅ Usando cloudscraper")
            else:
                print("  ⚠️  Usando requests padrão")
            
            print("  🔄 Estabelecendo conexão...")
            initial_response = self.session.get(self.base_url, timeout=15)
            if initial_response.status_code == 200:
                print("  ✅ Conexão estabelecida")
            self._initialized = True
        except Exception as e:
            print(f"  ⚠️  Aviso na conexão: {e}")
            self._initialized = True
    
    def get_player_stats_from_match(self, match_url, team, opponent, date, location):
        """Extrai estatísticas de jogadores de um jogo específico"""
        try:
            if not match_url or '/matches/' not in match_url:
                return []
            
            time.sleep(3)  # Rate limiting
            response = self.session.get(match_url, timeout=20)
            
            if response.status_code == 429:
                print(f"    ⚠️  Rate limit (429). Aguardando 30 segundos...")
                time.sleep(30)
                response = self.session.get(match_url, timeout=20)
            
            response.raise_for_status()
            soup = BeautifulSoup(response.content, 'html.parser')
            
            player_stats = []
            all_tables = soup.find_all('table', {'id': re.compile(r'.*')})
            
            # Debug: listar tabelas encontradas
            table_ids = [t.get('id', 'N/A') for t in all_tables if t.get('id')]
            if table_ids:
                print(f"    🔍 Tabelas encontradas: {len(table_ids)} (primeiras 5: {table_ids[:5]})")
            
            # Encontrar tabelas summary
            summary_tables = []
            for table in all_tables:
                table_id = table.get('id', '').lower()
                if 'summary' in table_id and 'stats' in table_id:
                    summary_tables.append(table)
            
            # Se encontrou tabelas summary, usar baseado na posição (home geralmente é primeira, away é segunda)
            stats_table = None
            if summary_tables:
                if location == 'home' and len(summary_tables) > 0:
                    stats_table = summary_tables[0]
                elif location == 'away' and len(summary_tables) > 1:
                    stats_table = summary_tables[1]
                elif len(summary_tables) == 1:
                    stats_table = summary_tables[0]
            
            # Se não encontrou summary, tentar padrões antigos como fallback
            if not stats_table:
                patterns = [
                    f'stats_{location}_summary',
                    f'stats_{location}_players',
                    f'stats_{location}',
                ]
                
                for pattern in patterns:
                    stats_table = soup.find('table', {'id': pattern})
                    if stats_table:
                        break
            
            # Método alternativo: procurar por tabelas com estrutura de estatísticas de jogadores
            if not stats_table:
                for table in all_tables:
                    table_id = table.get('id', '').lower()
                    if 'stats' in table_id and 'summary' in table_id:
                        thead = table.find('thead')
                        if thead:
                            headers = [th.get_text(strip=True).lower() for th in thead.find_all('th')]
                            header_text = ' '.join(headers)
                            if 'player' in header_text and ('min' in header_text or 'goals' in header_text):
                                stats_table = table
                                break
            
            # Último fallback: usar qualquer tabela com player
            if not stats_table:
                candidate_tables = []
                for table in all_tables:
                    thead = table.find('thead')
                    if thead:
                        headers = [th.get_text(strip=True).lower() for th in thead.find_all('th')]
                        header_text = ' '.join(headers)
                        if ('player' in header_text and 
                            ('min' in header_text or 'goals' in header_text or 'assists' in header_text)):
                            candidate_tables.append(table)
                
                summary_candidates = [t for t in candidate_tables if 'summary' in t.get('id', '').lower()]
                if summary_candidates:
                    candidate_tables = summary_candidates
                
                if candidate_tables:
                    if location == 'home' and len(candidate_tables) > 0:
                        stats_table = candidate_tables[0]
                    elif location == 'away' and len(candidate_tables) > 1:
                        stats_table = candidate_tables[1]
                    elif len(candidate_tables) == 1:
                        stats_table = candidate_tables[0]
            
            if not stats_table:
                print(f"    ❌ Tabela de estatísticas não encontrada para {location} team")
                return []
            
            print(f"    ✓ Tabela encontrada ({stats_table.get('id', 'N/A')})")
            rows = stats_table.find_all('tr')[1:]  # Pular cabeçalho
            first_row_processed = False
            
            for row in rows:
                row_class = str(row.get('class', []))
                if 'thead' in row_class or 'spacer' in row_class:
                    continue
                
                cells = row.find_all(['td', 'th'])
                if len(cells) < 3:
                    continue
                
                # Extrair dados
                player_name = ""
                minutes = 0
                goals = 0
                assists = 0
                xg = 0.0
                xa = 0.0
                
                # Primeiro loop: buscar valores básicos
                for cell in cells:
                    data_stat = cell.get('data-stat', '').lower()
                    text = cell.get_text(strip=True)
                    
                    if data_stat == 'player':
                        player_name = text
                    elif data_stat == 'minutes':
                        try:
                            minutes = int(re.sub(r'[^\d]', '', text) or 0)
                        except:
                            minutes = 0
                    elif data_stat == 'goals':
                        try:
                            goals = int(re.sub(r'[^\d]', '', text) or 0)
                        except:
                            goals = 0
                    elif data_stat == 'assists':
                        try:
                            assists = int(re.sub(r'[^\d]', '', text) or 0)
                        except:
                            assists = 0
                    elif data_stat == 'xg':
                        try:
                            text_clean = text.replace(',', '.')
                            text_clean = re.sub(r'[^\d.]', '', text_clean)
                            if text_clean:
                                xg = float(text_clean)
                            else:
                                xg = 0.0
                        except (ValueError, AttributeError):
                            xg = 0.0
                
                # Buscar xA APENAS usando xg_assist (nome correto no fbref)
                # NÃO usar fallbacks que podem pegar valores errados
                xa = 0.0
                for cell in cells:
                    data_stat = cell.get('data-stat', '').lower()
                    
                    # APENAS usar xg_assist - método mais confiável
                    if data_stat == 'xg_assist':
                        try:
                            text = cell.get_text(strip=True)
                            # Converter vírgula para ponto
                            text_clean = text.replace(',', '.')
                            # Remover tudo exceto dígitos e ponto
                            text_clean = re.sub(r'[^\d.]', '', text_clean)
                            if text_clean:
                                xa = float(text_clean)
                            else:
                                xa = 0.0
                            break  # Parar assim que encontrar (só existe uma célula xg_assist)
                        except (ValueError, AttributeError, TypeError):
                            xa = 0.0
                            break
                
                # Fallback para nome
                if not player_name and len(cells) > 0:
                    player_name = cells[0].get_text(strip=True)
                
                # Filtrar linhas inválidas
                if not player_name or player_name in ['Player', '', 'Reserves', 'Team Total']:
                    continue
                
                # Filtrar subtotais
                player_name_lower = player_name.lower()
                is_subtotal = False
                
                if 'player' in player_name_lower and any(char.isdigit() for char in player_name):
                    is_subtotal = True
                
                if minutes > 120:
                    is_subtotal = True
                
                if re.match(r'^\d+\s+[Pp]layers?$', player_name.strip()):
                    is_subtotal = True
                
                if is_subtotal:
                    continue
                
                # Fallback para estatísticas básicas (mas NÃO para xA)
                if minutes == 0 and goals == 0 and assists == 0:
                    header_row = stats_table.find('thead')
                    if header_row:
                        headers = [th.get_text(strip=True) for th in header_row.find_all('th')]
                        try:
                            min_idx = next((i for i, h in enumerate(headers) if 'min' in h.lower()), -1)
                            gls_idx = next((i for i, h in enumerate(headers) if 'gls' in h.lower() or h.lower() == 'g'), -1)
                            ast_idx = next((i for i, h in enumerate(headers) if 'ast' in h.lower() or h.lower() == 'a'), -1)
                            xg_idx = next((i for i, h in enumerate(headers) if 'xg' in h.lower() and 'xag' not in h.lower()), -1)
                            
                            if min_idx >= 0 and min_idx < len(cells):
                                minutes_str = cells[min_idx].get_text(strip=True)
                                minutes = int(re.sub(r'[^\d]', '', minutes_str) or 0)
                            
                            if gls_idx >= 0 and gls_idx < len(cells):
                                goals = int(re.sub(r'[^\d]', '', cells[gls_idx].get_text(strip=True)) or 0)
                            
                            if ast_idx >= 0 and ast_idx < len(cells):
                                assists = int(re.sub(r'[^\d]', '', cells[ast_idx].get_text(strip=True)) or 0)
                            
                            if xg_idx >= 0 and xg_idx < len(cells):
                                xg_str = cells[xg_idx].get_text(strip=True)
                                try:
                                    xg_str_clean = xg_str.replace(',', '.')
                                    xg_str_clean = re.sub(r'[^\d.]', '', xg_str_clean)
                                    if xg_str_clean:
                                        xg = float(xg_str_clean)
                                except (ValueError, AttributeError):
                                    pass
                        except (ValueError, IndexError):
                            pass
                
                # Formatar com 4 casas decimais (usar f-string para garantir trailing zeros)
                xg_formatted = f"{float(xg):.4f}"
                xa_formatted = f"{float(xa):.4f}"
                
                # Criar registro
                stats = {
                    'Player': player_name,
                    'Team': team,
                    'Date': date,
                    'Opponent': opponent,
                    'Minutes': minutes,
                    'Goals': goals,
                    'Assists': assists,
                    'xG': xg_formatted,
                    'xA': xa_formatted,
                    'Confronto': f"{team}|{opponent}|{date.strftime('%Y-%m-%d')}",
                    'Location': location,
                    'adj': 0,
                    'Year': date.year,
                    'Month': date.month
                }
                player_stats.append(stats)
            
            return player_stats
            
        except Exception as e:
            print(f"    ❌ Erro ao extrair estatísticas: {e}")
            return []


def get_season_url(league_id, year, month):
    """Gera URL da temporada baseado no mês"""
    if month >= 8:
        season = f"{year}-{year+1}"
    else:
        season = f"{year-1}-{year}"
    
    # Padrão de URL: https://fbref.com/en/comps/{league_id}/{season}/schedule/{season}-{League-Name}-Scores-and-Fixtures
    return f"https://fbref.com/en/comps/{league_id}/{season}/schedule/{season}-Scores-and-Fixtures"


def process_schedule_table(soup, start_date, end_date, scraper, limit_games=None):
    """Processa tabela de jogos"""
    all_player_stats = []
    
    # Encontrar tabela de jogos
    table = soup.find('table', {'id': re.compile(r'sched.*')})
    if not table:
        tables = soup.find_all('table')
        for t in tables:
            if 'schedule' in str(t.get('id', '')).lower():
                table = t
                break
    
    if not table:
        print(f"  ⚠️  Tabela de jogos não encontrada")
        return []
    
    rows = table.find_all('tr')[1:]
    print(f"  Encontradas {len(rows)} linhas na tabela")
    
    matches_found = 0
    skipped_no_date = 0
    skipped_out_of_range = 0
    skipped_no_teams = 0
    
    for row_idx, row in enumerate(rows):
        cells = row.find_all(['td', 'th'])
        if len(cells) < 3:
            continue
        
        try:
            # Extrair data
            date_text = ""
            date_cell = None
            
            for cell in cells:
                date_attr = cell.get('data-date', '')
                if date_attr and re.match(r'\d{4}-\d{2}-\d{2}', date_attr):
                    date_text = date_attr
                    date_cell = cell
                    break
            
            if not date_text:
                for cell in cells[:10]:
                    text = cell.get_text(strip=True)
                    if re.match(r'^\d{4}-\d{2}-\d{2}$', text):
                        date_text = text
                        date_cell = cell
                        break
            
            if not date_text:
                skipped_no_date += 1
                continue
            
            match_date = pd.to_datetime(date_text, format='%Y-%m-%d', errors='coerce')
            if pd.isna(match_date):
                skipped_no_date += 1
                continue
            
            if match_date < start_date or match_date > end_date:
                skipped_out_of_range += 1
                continue
            
            # Verificar se tem placar (jogo foi jogado)
            has_score = False
            score_text = ""
            for cell in cells:
                text = cell.get_text(strip=True)
                if re.match(r'^\d+[\s\-:]\d+$', text) and len(text) <= 10:
                    if ':' not in text or (':' in text and int(text.split(':')[0]) > 23):
                        has_score = True
                        score_text = text
                        break
            
            if not has_score:
                for cell in cells:
                    data_stat = cell.get('data-stat', '')
                    if 'score' in data_stat.lower() and cell.get_text(strip=True):
                        has_score = True
                        score_text = cell.get_text(strip=True)
                        break
            
            if not has_score:
                continue
            
            # Extrair times
            home_team = ""
            away_team = ""
            
            for cell in cells:
                data_stat = cell.get('data-stat', '')
                text = cell.get_text(strip=True)
                
                if data_stat == 'home_team':
                    home_team = text
                elif data_stat == 'away_team':
                    away_team = text
            
            if not home_team and len(cells) > 4:
                home_team = cells[4].get_text(strip=True)
            if not away_team and len(cells) > 5:
                away_team = cells[5].get_text(strip=True)
            
            if not home_team or not away_team:
                # Por links
                team_links = []
                for cell in cells:
                    link = cell.find('a')
                    if link:
                        text = link.get_text(strip=True)
                        href = link.get('href', '')
                        if text and len(text) > 2 and '/squads/' in href:
                            if (not re.match(r'^\d+$', text) and 
                                not re.match(r'\d{4}-\d{2}-\d{2}', text) and
                                not re.match(r'\d{2}:\d{2}', text) and
                                text not in ['Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat', 'Sun']):
                                team_links.append(text)
                
                if team_links:
                    if not home_team and len(team_links) > 0:
                        home_team = team_links[0]
                    if not away_team and len(team_links) > 1:
                        away_team = team_links[1]
            
            if not home_team or not away_team:
                skipped_no_teams += 1
                continue
            
            # Buscar link do jogo
            match_link = None
            
            # Primeiro, tentar na célula da data
            if date_cell:
                link_cell = date_cell.find('a')
                if link_cell and link_cell.get('href'):
                    href = link_cell.get('href', '')
                    if '/matches/' in href and '/schedule/' not in href:
                        # Ignorar URLs que são apenas datas (sem ID do match)
                        if not re.match(r'.*/matches/\d{4}-\d{2}-\d{2}$', href) and not re.match(r'/matches/\d{4}-\d{2}-\d{2}$', href):
                            match_link = urljoin(scraper.base_url, href)
            
            # Se não encontrou, procurar em todas as células
            if not match_link:
                for cell in cells:
                    links = cell.find_all('a')
                    for link in links:
                        href = link.get('href', '')
                        if href and '/matches/' in href and '/schedule/' not in href:
                            # Ignorar URLs que são apenas datas (sem ID do match)
                            if not re.match(r'.*/matches/\d{4}-\d{2}-\d{2}$', href) and not re.match(r'/matches/\d{4}-\d{2}-\d{2}$', href):
                                match_link = urljoin(scraper.base_url, href)
                                if match_link.startswith('http'):
                                    break
                    if match_link:
                        break
            
            # Se encontrou apenas URL de data, tentar construir URL completa a partir do texto
            if not match_link or re.match(r'.*/matches/\d{4}-\d{2}-\d{2}$', match_link):
                # Tentar encontrar link do jogo através dos nomes dos times
                for cell in cells:
                    links = cell.find_all('a')
                    for link in links:
                        href = link.get('href', '')
                        text = link.get_text(strip=True)
                        # Procurar links que contenham nome do time e match
                        if href and '/matches/' in href and text:
                            # Se o texto do link corresponde a um dos times
                            if (text.lower() in home_team.lower() or text.lower() in away_team.lower() or
                                home_team.lower() in text.lower() or away_team.lower() in text.lower()):
                                if '/schedule/' not in href and not re.match(r'.*/matches/\d{4}-\d{2}-\d{2}$', href):
                                    match_link = urljoin(scraper.base_url, href)
                                    break
                    if match_link and not re.match(r'.*/matches/\d{4}-\d{2}-\d{2}$', match_link):
                        break
            
            # Se ainda não encontrou, pular este jogo
            if not match_link or re.match(r'.*/matches/\d{4}-\d{2}-\d{2}$', match_link):
                print(f"     ⚠️  URL parece inválida (apenas data): {match_link if match_link else 'N/A'}")
                print(f"     ℹ️  Tentando encontrar link correto...")
                # Última tentativa: procurar qualquer link que não seja apenas data
                for cell in cells:
                    links = cell.find_all('a')
                    for link in links:
                        href = link.get('href', '')
                        if href and '/matches/' in href and len(href) > 20:  # Links de matches geralmente são longos
                            if '/schedule/' not in href and not re.match(r'.*/matches/\d{4}-\d{2}-\d{2}$', href):
                                match_link = urljoin(scraper.base_url, href)
                                break
                    if match_link and not re.match(r'.*/matches/\d{4}-\d{2}-\d{2}$', match_link):
                        break
            
            if not match_link or re.match(r'.*/matches/\d{4}-\d{2}-\d{2}$', match_link):
                continue
            
            print(f"     🔗 URL: {match_link}")
            
            if limit_games and matches_found >= limit_games:
                print(f"  ⏸️  Limite de {limit_games} jogos atingido")
                break
            
            print(f"  📅 {match_date.strftime('%Y-%m-%d')}: {home_team} vs {away_team} (Placar: {score_text})")
            
            # Buscar estatísticas
            stats_home = scraper.get_player_stats_from_match(
                match_link, home_team, away_team, match_date, 'home'
            )
            all_player_stats.extend(stats_home)
            
            stats_away = scraper.get_player_stats_from_match(
                match_link, away_team, home_team, match_date, 'away'
            )
            all_player_stats.extend(stats_away)
            
            matches_found += 1
            print(f"     ✓ {len(stats_home) + len(stats_away)} jogadores processados")
            
            time.sleep(5)
            
        except Exception as e:
            print(f"     ❌ Erro ao processar linha: {e}")
            continue
    
    print(f"\n  📊 Estatísticas:")
    print(f"     ✓ Jogos processados: {matches_found}")
    print(f"     ⚠️  Sem data: {skipped_no_date}")
    print(f"     ⚠️  Fora do período: {skipped_out_of_range}")
    print(f"     ⚠️  Sem times: {skipped_no_teams}")
    print(f"  ✅ Total: {len(all_player_stats)} registros de jogadores")
    
    return all_player_stats


def scrape_period(league_id, league_name, start_date, end_date, scraper=None, limit_games=None):
    """Busca dados para um período específico"""
    if scraper is None:
        scraper = LeagueScraper()
    
    if not scraper._initialized:
        scraper._ensure_initialized()
    
    all_player_stats = []
    
    # Gerar lista de meses
    current_date = start_date
    months_to_process = []
    
    while current_date <= end_date:
        year_month = (current_date.year, current_date.month)
        if year_month not in months_to_process:
            months_to_process.append(year_month)
        
        if current_date.month == 12:
            current_date = current_date.replace(year=current_date.year + 1, month=1, day=1)
        else:
            current_date = current_date.replace(month=current_date.month + 1, day=1)
    
    print(f"\n📅 Período: {start_date.strftime('%Y-%m-%d')} até {end_date.strftime('%Y-%m-%d')}")
    print(f"📋 Meses a processar: {', '.join([f'{y}-{m:02d}' for y, m in months_to_process])}")
    
    for year, month in months_to_process:
        print(f"\n{'='*60}")
        print(f"Processando {league_name} - {year}-{month:02d}...")
        print(f"{'='*60}")
        
        try:
            url = get_season_url(league_id, year, month)
            print(f"  Acessando: {url}")
            
            response = scraper.session.get(url, timeout=15)
            response.raise_for_status()
            soup = BeautifulSoup(response.content, 'html.parser')
            
            month_stats = process_schedule_table(soup, start_date, end_date, scraper, limit_games)
            all_player_stats.extend(month_stats)
            
            if (year, month) != months_to_process[-1]:
                print(f"  ⏳ Aguardando 10 segundos...")
                time.sleep(10)
                
        except Exception as e:
            print(f"  ❌ Erro ao processar {year}-{month:02d}: {e}")
            continue
    
    return all_player_stats


def main():
    parser = argparse.ArgumentParser(
        description='Busca estatísticas de jogadores de múltiplas ligas (MINUTES, GOALS, ASSISTS, XG, XA)',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Ligas disponíveis:
  --liga laliga        La Liga (Espanha)
  --liga bundesliga    Bundesliga (Alemanha)
  --liga seriea        Serie A (Itália)
  --liga portugal      Primeira Liga (Portugal)
  --liga ligue1        Ligue 1 (França)
  --liga championship  Championship (Inglaterra - Série B)
  --liga premier       Premier League (Inglaterra - Série A)

Exemplos:
  # La Liga
  python buscar_estatisticas_multi_liga.py --liga laliga --inicio 2025-09-01 --fim 2025-09-30

  # Bundesliga
  python buscar_estatisticas_multi_liga.py --liga bundesliga --inicio 2025-09-01 --fim 2025-09-30

  # Serie A
  python buscar_estatisticas_multi_liga.py --liga seriea --inicio 2025-09-01 --fim 2025-09-30

  # Primeira Liga (Portugal)
  python buscar_estatisticas_multi_liga.py --liga portugal --inicio 2025-09-01 --fim 2025-09-30

  # Ligue 1
  python buscar_estatisticas_multi_liga.py --liga ligue1 --inicio 2025-09-01 --fim 2025-09-30

  # Championship (Inglaterra Série B)
  python buscar_estatisticas_multi_liga.py --liga championship --inicio 2025-09-01 --fim 2025-09-30

  # Teste com 1 jogo
  python buscar_estatisticas_multi_liga.py --liga laliga --inicio 2025-09-01 --fim 2025-09-30 --limit 1 --test
        """
    )
    
    parser.add_argument('--liga', type=str, required=True,
                       choices=['laliga', 'bundesliga', 'seriea', 'portugal', 'ligue1', 'championship', 'premier'],
                       help='Liga a buscar dados')
    parser.add_argument('--inicio', type=str, required=True,
                       help='Data de início (YYYY-MM-DD)')
    parser.add_argument('--fim', type=str, required=True,
                       help='Data de fim (YYYY-MM-DD)')
    parser.add_argument('--output', type=str, default=None,
                       help='Arquivo Excel de saída (padrão: {liga}_{data}.xlsx)')
    parser.add_argument('--limit', type=int, default=None,
                       help='Limitar número de jogos')
    parser.add_argument('--test', action='store_true',
                       help='Modo teste - não salva arquivo')
    
    args = parser.parse_args()
    
    league_info = LEAGUE_IDS[args.liga]
    league_id = league_info['id']
    league_name = league_info['name']
    country = league_info['country']
    
    print("="*60)
    print(f"BUSCADOR DE ESTATÍSTICAS - {league_name.upper()}")
    print(f"País: {country}")
    print("="*60)
    print("\nEstatísticas buscadas: MINUTES, GOALS, ASSISTS, XG, XA")
    
    if args.test:
        print("🧪 MODO TESTE - Nenhuma alteração será salva")
    
    # Parsear datas
    try:
        start_date = pd.to_datetime(args.inicio)
        end_date = pd.to_datetime(args.fim)
    except:
        print("❌ Erro: Formato de data inválido. Use YYYY-MM-DD")
        return
    
    if start_date > end_date:
        print("❌ Erro: Data de início deve ser anterior à data de fim")
        return
    
    # Nome do arquivo de saída
    if not args.output:
        league_slug = args.liga.replace(' ', '_')
        args.output = f"{league_slug}_{start_date.strftime('%Y%m%d')}_{end_date.strftime('%Y%m%d')}.xlsx"
    
    print(f"\n📅 Período: {start_date.strftime('%Y-%m-%d')} até {end_date.strftime('%Y-%m-%d')}")
    print(f"📁 Arquivo de saída: {args.output}")
    
    if not args.test:
        resposta = input("\n⚠️  Continuar com a busca? (s/n): ").strip().lower()
        if resposta not in ['s', 'sim', 'y', 'yes']:
            print("Operação cancelada.")
            return
    
    # Inicializar scraper
    print("\n🔧 Inicializando scraper...")
    scraper = LeagueScraper()
    scraper._ensure_initialized()
    
    # Buscar dados
    print("\n🚀 Iniciando busca...")
    all_data = scrape_period(league_id, league_name, start_date, end_date, scraper, limit_games=args.limit)
    
    if not all_data:
        print("\n⚠️  Nenhum dado foi encontrado.")
        return
    
    # Converter para DataFrame
    print(f"\n📊 Processando {len(all_data)} registros coletados...")
    new_df = pd.DataFrame(all_data)
    
    # Remover duplicatas
    before_dedup = len(new_df)
    new_df = new_df.drop_duplicates(subset=['Player', 'Team', 'Date', 'Opponent'], keep='last')
    duplicates_removed = before_dedup - len(new_df)
    
    if duplicates_removed > 0:
        print(f"  🧹 Removidas {duplicates_removed} duplicatas")
    
    # Ordenar por data
    new_df = new_df.sort_values('Date')
    
    if args.test:
        print(f"\n{'='*60}")
        print("🧪 MODO TESTE - Resultados:")
        print(f"{'='*60}")
        print(f"📈 Total de registros: {len(new_df)}")
        if len(new_df) > 0:
            print(f"📅 Data mínima: {new_df['Date'].min().strftime('%Y-%m-%d')}")
            print(f"📅 Data máxima: {new_df['Date'].max().strftime('%Y-%m-%d')}")
            print(f"\nPrimeiras 5 linhas:")
            print(new_df.head(5).to_string())
        print(f"\n💡 Execute sem --test para salvar a planilha")
        print(f"{'='*60}")
    else:
        # Salvar planilha
        import os
        from pathlib import Path
        from openpyxl import load_workbook
        
        output_path = Path(args.output)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        
        new_df.to_excel(args.output, index=False)
        
        # Formatar xG e xA com 4 casas decimais
        try:
            wb = load_workbook(args.output)
            ws = wb.active
            
            header_row = 1
            xg_col = None
            xa_col = None
            
            for col_idx, cell in enumerate(ws[header_row], 1):
                if cell.value == 'xG':
                    xg_col = col_idx
                elif cell.value == 'xA':
                    xa_col = col_idx
            
            if xg_col:
                for row in range(2, ws.max_row + 1):
                    cell = ws.cell(row=row, column=xg_col)
                    if cell.value is not None:
                        try:
                            cell.value = float(cell.value)
                            cell.number_format = "0.0000"
                        except:
                            pass
            
            if xa_col:
                for row in range(2, ws.max_row + 1):
                    cell = ws.cell(row=row, column=xa_col)
                    if cell.value is not None:
                        try:
                            cell.value = float(cell.value)
                            cell.number_format = "0.0000"
                        except:
                            pass
            
            wb.save(args.output)
            wb.close()
        except Exception as e:
            print(f"  ⚠️  Aviso: Não foi possível formatar decimais: {e}")
        
        print(f"\n{'='*60}")
        print("✅ PLANILHA SALVA COM SUCESSO!")
        print(f"{'='*60}")
        print(f"📈 Total de registros: {len(new_df)}")
        if len(new_df) > 0:
            print(f"📅 Data mínima: {new_df['Date'].min().strftime('%Y-%m-%d')}")
            print(f"📅 Data máxima: {new_df['Date'].max().strftime('%Y-%m-%d')}")
        print(f"📁 Arquivo salvo: {os.path.abspath(args.output)}")
        print(f"{'='*60}")


if __name__ == "__main__":
    main()

